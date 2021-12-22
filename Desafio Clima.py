from openpyxl import load_workbook

from openpyxl import Workbook

from openpyxl.styles import numbers

from os import listdir

import string

import re

import time
##

def unique(list1):
    # intilize a null list
    unique_list = []

    # traverse for all elements
    for x in list1:
        # check if exists in unique_list or not
        if x not in unique_list:
            unique_list.append(x)
    return(unique_list)


def lista_str_num_inv(letra):
#faz as combinacoes de letras como no excel
    abc = string.ascii_uppercase[:27]
    abc_ind = list()

    tamanho = len(letra)
    cont = 0
    ciclo = 0
    pol = 0

    cont_letra = 0
    for i in range(tamanho,0,-1):
        pol = pol + (abc.find(letra[cont_letra])+1)*26**(i-1)
        cont_letra = cont_letra + 1

    return(pol)



def leitura_conj_lista(Aba, coluna, linha, limite_linha, campos_vazio):

    lista = list()

    col = lista_str_num_inv(coluna)

    cont = linha

    for value in Aba.iter_rows(min_row=linha, max_row = limite_linha, min_col=col, max_col=col,
                                                            values_only=True):
        if list(value)[0] != None:

            lista.append(cont)

        cont = cont + 1

    return (lista)



'''def leitura_conj_lista(Aba, coluna,linha, limite_linha,campos_vazios):
#retorna as linhas onde tem dados
#cont_vazio -> Quantas linhas vazias são toleradas até parar de contar

    linhas_com_dados = list()
    cont_vazio = 0
    for i in range(linha, limite_linha):
        if Aba[coluna+str(i)].value is not None:
            linhas_com_dados.append(i)
            cont_vazio = 0
        else:
            cont_vazio = cont_vazio + 1

        if cont_vazio >= campos_vazios:
            break


    return(linhas_com_dados)'''


def lista_str_num(indice):
#faz as combinacoes de letras como no excel
    abc = string.ascii_uppercase[:27]
    abc_ind = list()

    for i in range(len(abc)):

        abc_ind.append(abc[i])

    if indice <= 26:
        return(abc_ind[indice-1])
    else:
        k = (indice-1)//26-1
        l = indice%26-1
        return(abc_ind[k]+abc_ind[l])


def dic_listas_GHG(Aba):
    dic_listas = dict()

    i=1
    cont_vazio = 0
    while cont_vazio < 10:
        if Aba[lista_str_num(i)+str(1)].value is not None:
            dic_listas[Aba[lista_str_num(i)+str(1)].value] = lista_str_num(i)
            cont_vazio = 0
        else:
            cont_vazio = cont_vazio + 1
        i=i+1


def listas_GHG(Aba, dic_listas, lista):

   col_lista = dic_listas[lista]

   lista2 = list()

   i = 3
   cont_vazio = 0
   while cont_vazio < 1:
       if Aba[col_lista + str(i)].value is not None:
           lista2.append(Aba[col_lista + str(i)].value)
           cont_vazio = 0
           i=i+1
       else:
           cont_vazio = cont_vazio+1

    return(lista2)



def leitura_conj_dict(Aba, coluna, coluna_chave,linha, limite_linha):

    linhas_com_dados = list()
    chaves = list()
    dic_chaves = dict()
    for i in range(linha, limite_linha):
        if(Aba[coluna+str(i)]).value is not None:
            linhas_com_dados.append(i)
            chaves.append(Aba[coluna_chave+str(i)].value)

    for chave in dic_chaves:

        if chave not in dic_chaves:
            dic_chaves[chave] = 1
        else:
            dic_plans[chave] = dic_plans[chave] + 1


    return(linhas_com_dados)


#Buscar o nome das abas da planilha do GHG daquele ano

pasta_Plan_atual = "C:\\Users\\Erick\\Desktop\\PHS\\Planilha Clima Vigente"

Plan_atual = listdir(pasta_Plan_atual)

end_plan_atual = pasta_Plan_atual + "\\" + Plan_atual[0]

Plan_atual = load_workbook(end_plan_atual, read_only=True, data_only=True)

lista_nomes_planGHG = Plan_atual.sheetnames[4:]




#criando a planilha de extração

#pasta_BD = "C:\\Users\\Erick\\Desktop\\PHS\\Desafio Clima_Codigos\\Planilhas BD_Teste"
pasta_BD = "C:\\Users\\Erick\\Desktop\\PHS\\Desafio Clima_Codigos\\GHG\\Inventários-GHG"
pasta_saida = "C:\\Users\\Erick\\Desktop\\PHS\\Desafio Clima_Codigos\\Planilha Saída"

filename = pasta_saida + "\\Planilha_Saída.xlsx"
plan_saida = Workbook()

cont = 0

for i in lista_nomes_planGHG:

    plan_saida.create_sheet(i, cont)
    cont = cont +1


# Criando duas abas para as emissoes fugitivas
tit_aux = plan_saida.worksheets[2].title
plan_saida.worksheets[2].title = plan_saida.worksheets[2].title + "1"
plan_saida.create_sheet(tit_aux+"2", 3)

# Criando duas abas para Resíduos sólidos
tit_aux = plan_saida.worksheets[7].title
plan_saida.worksheets[7].title = plan_saida.worksheets[7].title + "1"
plan_saida.create_sheet(tit_aux+"2", 8)

# Criando tres abas para En. elétrica (localizacao)

tit_aux = plan_saida.worksheets[10].title
plan_saida.worksheets[10].title = plan_saida.worksheets[10].title + "1"
plan_saida.create_sheet(tit_aux+"2", 11)
plan_saida.create_sheet(tit_aux+"3", 12)

# Criando duas abas para Resíduos Sólidos da Operação

tit_aux = plan_saida.worksheets[19].title
plan_saida.worksheets[19].title = plan_saida.worksheets[19].title + "1"
plan_saida.create_sheet(tit_aux+"2", 20)

# Criando duas abas para Viagens a Negócios

tit_aux = plan_saida.worksheets[22].title
plan_saida.worksheets[22].title = plan_saida.worksheets[22].title + "1"
plan_saida.create_sheet(tit_aux+"2", 23)

# Criando quatro abas para Deslocamento casa-trabalho

tit_aux = plan_saida.worksheets[24].title
plan_saida.worksheets[24].title = plan_saida.worksheets[24].title + "1"
plan_saida.create_sheet(tit_aux+"2", 25)
plan_saida.create_sheet(tit_aux+"3", 26)
plan_saida.create_sheet(tit_aux+"4", 27)




Plans = listdir(pasta_BD)
Plans_CNPJ_data = list()
Plans_CNPJ = list()

for planilha in Plans:

    Plans_CNPJ_data.append(re.findall('(^.+)_' ,planilha))
    Plans_CNPJ.append(re.findall('(^.+?)_' ,planilha))

dic_plans = dict()

for p in range(len(Plans_CNPJ)):

    if Plans_CNPJ[p][0] not in dic_plans:
        dic_plans[Plans_CNPJ[p][0]]=1
    else:
        dic_plans[Plans_CNPJ[p][0]] = dic_plans[Plans_CNPJ[p][0]] + 1


# Nome das colunas

#Combustão estacionária

Aba = plan_saida["Combustão estacionária"]
Comb_est =  Plan_atual["Combustão estacionária"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Comb_est["C" + str(42)].value # Descricao da fonte
Aba["F" + str(1)] = Comb_est["D" + str(42)].value # Combustivel
Aba["G" + str(1)] = Comb_est["E" + str(42)].value # Quantidade
Aba["H" + str(1)] = Comb_est["F" + str(42)].value # Unidade
Aba["I" + str(1)] = Comb_est["Q" + str(43)].value # Quantidade CO2 Fossil
Aba["J" + str(1)] = Comb_est["R" + str(43)].value # Quantidade CH4 Fossil
Aba["K" + str(1)] = Comb_est["S" + str(43)].value # Quantidade N2O Fossil
Aba["L" + str(1)] = Comb_est["T" + str(43)].value # Quantidade CO2 Biocomb
Aba["M" + str(1)] = Comb_est["U" + str(43)].value # Quantidade CH4 Biocomb
Aba["N" + str(1)] = Comb_est["V" + str(43)].value # Quantidade N2O Biocomb
Aba["O" + str(1)] = Comb_est["W" + str(42)].value # Emissoes fosseis
Aba["P" + str(1)] = Comb_est["X" + str(42)].value # Emissoes biogenicas

# Combustão móvel

Aba = plan_saida["Combustão móvel"]
Comb_mov =  Plan_atual["Combustão móvel"]

#As linhas abaixo extraem os cabeçalhos das linhas da Tabela 1 por isso estão comentadas
'''Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Comb_mov["C" + str(41)].value #Descricao da frota
Aba["F" + str(1)] = Comb_mov["D" + str(41)].value #Tipo de frota
Aba["G" + str(1)] = Comb_mov["E" + str(41)].value #Ano da frota
Aba["H" + str(1)] = Comb_mov["R" + str(42)].value #Consumo anual
Aba["I" + str(1)] = Comb_mov["AZ" + str(41)].value #Emissoes CO2 fossil
Aba["J" + str(1)] = Comb_mov["BA" + str(41)].value #Emissoes CH4 fossil
Aba["K" + str(1)] = Comb_mov["BB" + str(41)].value #Emissoes N2O
Aba["L" + str(1)] = Comb_mov["BC" + str(41)].value #Emissoes totais eq
Aba["M" + str(1)] = Comb_mov["BD" + str(41)].value #Emissoes CO2 biog'''

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Comb_mov["C" + str(203)].value
Aba["F" + str(1)] = Comb_mov["D" + str(203)].value
Aba["G" + str(1)] = Comb_mov["R" + str(204)].value
Aba["H" + str(1)] = Comb_mov["BB" + str(203)].value
Aba["I" + str(1)] = Comb_mov["BC" + str(203)].value
Aba["J" + str(1)] = Comb_mov["BD" + str(203)].value
Aba["K" + str(1)] = Comb_mov["BE" + str(203)].value
Aba["L" + str(1)] = Comb_mov["BF" + str(203)].value



# Emissoes fugitivas

Em_fug =  Plan_atual["Emissões fugitivas"]
Aba = plan_saida["Emissões fugitivas1"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Em_fug["B" + str(58)].value #Registro da fonte
Aba["F" + str(1)] = Em_fug["C" + str(58)].value #Gas ou composto
Aba["G" + str(1)] = Em_fug["E" + str(58)].value + " " + Em_fug["E" + str(59)].value + " " + Em_fug["E" + str(60)].value #Unidades novas carga
Aba["H" + str(1)] = Em_fug["E" + str(58)].value + " " + Em_fug["F" + str(59)].value + " " + Em_fug["F" + str(60)].value #Unidades novas capacidade
Aba["I" + str(1)] = Em_fug["G" + str(58)].value + " " + Em_fug["G" + str(59)].value + " " + Em_fug["G" + str(60)].value #Unidades existentes recarga
Aba["J" + str(1)] = Em_fug["H" + str(58)].value + " " + Em_fug["H" + str(59)].value + " " + Em_fug["H" + str(60)].value #Unidades dispensadas capacidade
Aba["K" + str(1)] = Em_fug["H" + str(58)].value + " " + Em_fug["I" + str(59)].value + " " + Em_fug["I" + str(60)].value #Unidades dispensadas recuperada
Aba["L" + str(1)] = Em_fug["J" + str(58)].value + " " + Em_fug["J" + str(59)].value + " " + Em_fug["J" + str(60)].value #Emissoes CO2eq


Aba = plan_saida["Emissões fugitivas2"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Em_fug["B" + str(235)].value #Registro da fonte
Aba["F" + str(1)] = Em_fug["C" + str(235)].value #Descricao da fonte
Aba["G" + str(1)] = Em_fug["D" + str(235)].value #Gas de Efeito Estufa
Aba["H" + str(1)] = Em_fug["E" + str(235)].value #Emissoes (kg GEE)
Aba["I" + str(1)] = Em_fug["G" + str(235)].value #Emissoes em CO2e (t)

# Residuos solidos

Res_sol = Plan_atual["Resíduos sólidos"]
Aba = plan_saida["Resíduos sólidos1"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Res_sol["C"+ str(135)].value #Massa de resíduo destinado à compostagem
Aba["F" + str(1)] = Res_sol["C" + str(154)].value + Res_sol["D" + str(154)].value #Emissoes de CH4 por compostagem
Aba["G" + str(1)] = Res_sol["C" + str(155)].value + Res_sol["D" + str(155)].value #Emissoes de N2O por compostagem
Aba["H" + str(1)] = Res_sol["C" + str(156)].value + Res_sol["D" + str(156)].value #Emissoes de tCO2 por compostagem
Aba["I" + str(1)] = Res_sol["C" + str(157)].value + Res_sol["D" + str(157)].value #Emissoes de CO2 biog

Aba = plan_saida["Resíduos sólidos2"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Res_sol["C"+str(168)].value #Tipo de residuo incinerado
Aba["F" + str(1)] = Res_sol["D"+str(168)].value #Emissões CO2 (t)
Aba["G" + str(1)] = Res_sol["E"+str(168)].value #Emissoes CH4 (t)
Aba["H" + str(1)] = Res_sol["F"+str(168)].value #Emissoes N2O (t)
Aba["I" + str(1)] = Res_sol["G"+str(168)].value #Emissoes CO2eq (t)
Aba["J" + str(1)] = Res_sol["H"+str(168)].value #Emissoes CO2eq (t)

# Efluentes

Efluentes = Plan_atual["Efluentes"]
Aba = plan_saida["Efluentes"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Efluentes["C" + str(34)].value
Aba["F" + str(1)] = Efluentes["C" + str(39)].value
Aba["G" + str(1)] = Efluentes["C" + str(46)].value
Aba["H" + str(1)] = Efluentes["E" + str(45)].value # Unidade
Aba["I" + str(1)] = Efluentes["C" + str(52)].value # Quantidade N2
Aba["J" + str(1)] = Efluentes["C" + str(53)].value # Fator emissão N2O
Aba["K" + str(1)] = Efluentes["C" + str(54)].value # N removido
Aba["L" + str(1)] = Efluentes["C" + str(58)].value # Tipo de tratamento
Aba["M" + str(1)] = Efluentes["C" + str(59)].value # Fator de conversão de metano
Aba["N" + str(1)] = Efluentes["C" + str(64)].value # CH4 recuperado
Aba["O" + str(1)] = Efluentes["C" + str(67)].value # Destino biogas

# Energia elétrica
#1
En_elet = Plan_atual["En. elétrica (localização)"]
Aba = plan_saida["En. elétrica (localização)1"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = En_elet["C" + str(37)].value
Aba["F" + str(1)] = En_elet["Q" + str(37)].value
Aba["G" + str(1)] = En_elet["AD" + str(38)].value

#2
Aba = plan_saida["En. elétrica (localização)2"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = En_elet["C" + str(102)].value
Aba["F" + str(1)] = En_elet["Q" + str(102)].value
Aba["G" + str(1)] = En_elet["AD" + str(102)].value

#3
Aba = plan_saida["En. elétrica (localização)3"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = En_elet["C" + str(134)].value
Aba["F" + str(1)] = En_elet["D" + str(134)].value
Aba["G" + str(1)] = En_elet["F" + str(134)].value
Aba["H" + str(1)] = En_elet["H" + str(134)].value
Aba["I" + str(1)] = En_elet["J" + str(134)].value
Aba["J" + str(1)] = En_elet["L" + str(134)].value

# Energia elétrica (escolha de compra)

En_elet_2 = Plan_atual["En. elétrica(escolha de compra)"]
Aba = plan_saida["En. elétrica(escolha de compra)"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = En_elet_2["C" + str(39)].value
Aba["F" + str(1)] = En_elet_2["D" + str(39)].value
Aba["G" + str(1)] = En_elet_2["E" + str(39)].value
Aba["H" + str(1)] = En_elet_2["F" + str(39)].value
Aba["I" + str(1)] = En_elet_2["G" + str(39)].value
Aba["J" + str(1)] = En_elet_2["AC" + str(39)].value
Aba["K" + str(1)] = En_elet_2["AD" + str(39)].value
Aba["L" + str(1)] = En_elet_2["AE" + str(39)].value
Aba["M" + str(1)] = En_elet_2["AF" + str(39)].value
Aba["N" + str(1)] = En_elet_2["AG" + str(39)].value
Aba["O" + str(1)] = En_elet_2["AH" + str(39)].value

# Transporte e Distribuição (Upstream)

nome_aba = "Transp.& Distribuição(Upstream)"

cont_aux = 0
for nome in Plan_atual.sheetnames:
    if nome != nome_aba:
        cont_aux = cont_aux + 1
    else:
        break

cont_aux2 = 0
for nome in plan_saida.sheetnames:
    if nome != nome_aba:
        cont_aux2 = cont_aux2 + 1
    else:
        break

T_D_Up = Plan_atual.worksheets[cont_aux]
Aba = plan_saida.worksheets[cont_aux2]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = T_D_Up["C" + str(536)].value #Descrição da fonte
Aba["F" + str(1)] = T_D_Up["L" + str(536)].value #Tipo da frota de veículos
Aba["G" + str(1)] = T_D_Up["M" + str(536)].value #Ano da frota
Aba["H" + str(1)] = T_D_Up["AH" + str(537)].value #Distância anual (km)
Aba["I" + str(1)] = T_D_Up["CE" + str(536)].value #Emissões CO2 Fóssil (t)
Aba["J" + str(1)] = T_D_Up["CF" + str(536)].value #Emissões CH4 (t)
Aba["K" + str(1)] = T_D_Up["CG" + str(536)].value #Emissões N2O (t)
Aba["L" + str(1)] = T_D_Up["CH" + str(536)].value #Emissões totais (t CO2e)
Aba["M" + str(1)] = T_D_Up["CI" + str(536)].value #Emissões CO2 biog

# Residuos Sólidos da Operação

Res_sol_op = Plan_atual["Resíduos sólidos da operação"]
Aba = plan_saida["Resíduos sólidos da operação1"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Res_sol_op["C" + str(54)].value #Estado
Aba["F" + str(1)] = Res_sol_op["C" + str(55)].value #Municipio
Aba["G" + str(1)] = Res_sol_op["C" + str(67)].value + Res_sol_op["D" + str(67)].value
Aba["H" + str(1)] = Res_sol_op["C" + str(103)].value
Aba["I" + str(1)] = Res_sol_op["C" + str(114)].value
Aba["J" + str(1)] = Res_sol_op["C" + str(125)].value + Res_sol_op["D" + str(125)].value
Aba["K" + str(1)] = Res_sol_op["C" + str(126)].value + Res_sol_op["D" + str(126)].value
Aba["L" + str(1)] = Res_sol_op["C" + str(127)].value + Res_sol_op["D" + str(127)].value

Aba = plan_saida["Resíduos sólidos da operação2"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Res_sol_op["C" + str(168)].value
Aba["F" + str(1)] = Res_sol_op["D" + str(168)].value
Aba["G" + str(1)] = Res_sol_op["E" + str(168)].value
Aba["H" + str(1)] = Res_sol_op["F" + str(168)].value
Aba["I" + str(1)] = Res_sol_op["G" + str(168)].value
Aba["J" + str(1)] = Res_sol_op["H" + str(168)].value

# Efluentes gerados na operação

Eflu_op = Plan_atual["Efluentes gerados na operação"]
Aba = plan_saida["Efluentes gerados na operação"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Eflu_op["C" + str(40)].value
Aba["F" + str(1)] = Eflu_op["C" + str(45)].value
Aba["G" + str(1)] = Eflu_op["C" + str(52)].value
Aba["H" + str(1)] = Eflu_op["E" + str(51)].value
Aba["I" + str(1)] = Eflu_op["C" + str(58)].value
Aba["J" + str(1)] = Eflu_op["C" + str(59)].value
Aba["K" + str(1)] = Eflu_op["C" + str(60)].value
Aba["L" + str(1)] = Eflu_op["C" + str(64)].value
Aba["M" + str(1)] = Eflu_op["C" + str(65)].value
Aba["N" + str(1)] = Eflu_op["C" + str(70)].value
Aba["O" + str(1)] = Eflu_op["C" + str(79)].value
Aba["P" + str(1)] = Eflu_op["C" + str(86)].value
Aba["Q" + str(1)] = Eflu_op["D" + str(85)].value
Aba["R" + str(1)] = Eflu_op["C" + str(92)].value
Aba["S" + str(1)] = Eflu_op["C" + str(93)].value
Aba["T" + str(1)] = Eflu_op["C" + str(94)].value
Aba["U" + str(1)] = Eflu_op["C" + str(98)].value
Aba["V" + str(1)] = Eflu_op["C" + str(99)].value
Aba["W" + str(1)] = Eflu_op["C" + str(104)].value
Aba["X" + str(1)] = Eflu_op["C" + str(110)].value
Aba["Y" + str(1)] = Eflu_op["C" + str(111)].value
Aba["Z" + str(1)] = Eflu_op["C" + str(112)].value
Aba["AA" + str(1)] = Eflu_op["C" + str(113)].value

# Viagens a negócios

#1

Viagens = Plan_atual["Viagens a Negócios"]
Aba = plan_saida["Viagens a Negócios1"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Viagens["W" + str(45)].value
Aba["F" + str(1)] = Viagens["X" + str(45)].value
Aba["G" + str(1)] = Viagens["Y" + str(45)].value
Aba["H" + str(1)] = Viagens["AC" + str(45)].value
Aba["I" + str(1)] = Viagens["AD" + str(45)].value
Aba["J" + str(1)] = Viagens["AE" + str(45)].value
Aba["K" + str(1)] = Viagens["AF" + str(45)].value

#2

Viagens = Plan_atual["Viagens a Negócios"]
Aba = plan_saida["Viagens a Negócios2"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Viagens["C" + str(1098)].value
Aba["F" + str(1)] = Viagens["L" + str(1098)].value
Aba["G" + str(1)] = Viagens["M" + str(1098)].value
Aba["H" + str(1)] = Viagens["AH" + str(1099)].value
Aba["I" + str(1)] = Viagens["CE" + str(1098)].value
Aba["J" + str(1)] = Viagens["CF" + str(1098)].value
Aba["K" + str(1)] = Viagens["CG" + str(1098)].value
Aba["L" + str(1)] = Viagens["CH" + str(1098)].value
Aba["M" + str(1)] = Viagens["CI" + str(1098)].value

# Deslocamento casa-trabalho

#1

Desloc_CT = Plan_atual["Deslocamento casa-trabalho"]
Aba = plan_saida["Deslocamento casa-trabalho1"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Desloc_CT["D" + str(37)].value
Aba["F" + str(1)] = Desloc_CT["F" + str(37)].value
Aba["G" + str(1)] = Desloc_CT["G" + str(37)].value
Aba["H" + str(1)] = Desloc_CT["K" + str(37)].value
Aba["I" + str(1)] = Desloc_CT["N" + str(37)].value

#2

Aba = plan_saida["Deslocamento casa-trabalho2"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Desloc_CT["D" + str(149)].value
Aba["F" + str(1)] = Desloc_CT["F" + str(149)].value
Aba["G" + str(1)] = Desloc_CT["G" + str(149)].value
Aba["H" + str(1)] = Desloc_CT["W" + str(149)].value
Aba["I" + str(1)] = Desloc_CT["X" + str(149)].value

#3

Aba = plan_saida["Deslocamento casa-trabalho3"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Desloc_CT["D" + str(262)].value
Aba["F" + str(1)] = Desloc_CT["F" + str(262)].value
Aba["G" + str(1)] = Desloc_CT["G" + str(262)].value
Aba["H" + str(1)] = Desloc_CT["K" + str(263)].value
Aba["I" + str(1)] = Desloc_CT["L" + str(263)].value
Aba["J" + str(1)] = Desloc_CT["M" + str(263)].value
Aba["K" + str(1)] = Desloc_CT["N" + str(262)].value


#4

Aba = plan_saida["Deslocamento casa-trabalho4"]

Aba["A" + str(1)] = "CNPJ"
Aba["B" + str(1)] = "Nome da Instituição"
Aba["C" + str(1)] = "Ano"
Aba["D" + str(1)] = "Nome do responsável"
Aba["E" + str(1)] = Desloc_CT["D" + str(468)].value
Aba["F" + str(1)] = Desloc_CT["F" + str(468)].value
Aba["G" + str(1)] = Desloc_CT["T" + str(468)].value
Aba["H" + str(1)] = Desloc_CT["BC" + str(468)].value
Aba["I" + str(1)] = Desloc_CT["BD" + str(468)].value
Aba["J" + str(1)] = Desloc_CT["BE" + str(468)].value
Aba["K" + str(1)] = Desloc_CT["BF" + str(468)].value
Aba["L" + str(1)] = Desloc_CT["BG" + str(468)].value




print (time.strftime("%H:%M:%S"))
# _#################################### Combustão Estacionária ###############################################
cont = 1
for planilha,i in zip(Plans,range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only = True, data_only= True)


    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Combustão estacionária"]

    Aba["A"+str(cont+1)] = int(re.findall('(^.+?)_' ,Plans[i])[0])
    Aba["A" + str(cont+1)].number_format =  numbers.builtin_format_code(1)
    Aba["B"+str(cont+1)] = nome_org.value
    Aba["C"+str(cont+1)] = ano.value
    Aba["D"+str(cont+1)] = nome_resp.value

    Comb_est = workbook["Combustão estacionária"]

    lista_aux_Comb_est = leitura_conj_lista(Comb_est, "D", 46, 145,10)

    '''if lista_aux_Comb_est == []:
        for i in string.ascii_uppercase[4:15]:
            if i != "H":
                Aba[i + str(cont+1)] = 0'''

    for j,k in zip(lista_aux_Comb_est, range(1,(len(lista_aux_Comb_est)+1))):

        Desc_Fonte = Comb_est["C" + str(j)]
        Comb_Util = Comb_est["D" + str(j)]
        Quant_cons = Comb_est["E" + str(j)]
        Unidade = Comb_est["F" + str(j)]
        CO2_Foss = Comb_est["Q" + str(j)]
        CH4_Foss = Comb_est["R" + str(j)]
        N2O_Foss = Comb_est["S" + str(j)]
        CO2_Biocomb = Comb_est["T" + str(j)]
        CH4_Biocomb = Comb_est["U" + str(j)]
        N2O_Biocomb = Comb_est["V" + str(j)]
        Emissao_Foss = Comb_est["W" + str(j)]
        Emissao_biog = Comb_est["X" + str(j)]

        Aba["E"+str(cont+k)] = Desc_Fonte.value
        Aba["F"+str(cont+k)] = Comb_Util.value
        Aba["G" + str(cont+k)] = Quant_cons.value
        Aba["H" + str(cont+k)] = Unidade.value

        if Quant_cons.value is not None:
            Aba["I" + str(cont+k)] = float(CO2_Foss.value)
            Aba["J" + str(cont + k)] = float(CH4_Foss.value)
            Aba["K" + str(cont + k)] = float(N2O_Foss.value)
            Aba["L" + str(cont + k)] = float(CO2_Biocomb.value)
            Aba["M" + str(cont + k)] = float(CH4_Biocomb.value)
            Aba["N" + str(cont + k)] = float(N2O_Biocomb.value)
            Aba["O" + str(cont + k)] = float(Emissao_Foss.value)


        else:
            Aba["I" + str(cont+k)] = CO2_Foss.value
            Aba["J" + str(cont + k)] = CH4_Foss.value
            Aba["K" + str(cont + k)] = N2O_Foss.value
            Aba["L" + str(cont + k)] = CO2_Biocomb.value
            Aba["M" + str(cont + k)] = CH4_Biocomb.value
            Aba["N" + str(cont + k)] = N2O_Biocomb.value
            Aba["O" + str(cont + k)] = Emissao_Foss.value

        if Emissao_biog.value is not None:
            Aba["P" + str(cont + k)] = float(Emissao_biog.value)
        else:
            Aba["P" + str(cont + k)] = Emissao_biog.value

        if Emissao_Foss.value == "0":
            Aba["O" + str(cont + k)] = 0

        if Emissao_biog.value == "0":
            Aba["P" + str(cont + k)] = 0

        Aba["A" + str(cont+k)] = int(re.findall('(^.+?)_' ,Plans[i])[0])
        Aba["A" + str(cont + k)].number_format =  numbers.builtin_format_code(1)
        Aba["B" + str(cont+k)] = nome_org.value
        Aba["C" + str(cont+k)] = ano.value
        Aba["D" + str(cont+k)] = nome_resp.value

    cont = cont+k

print("Fim do processamento de Combustão Estacionária")



#_################################################# Combustão Móvel ##############################################################

cont = 1
for planilha, i in zip(Plans, range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Combustão móvel"]

    Aba["A" + str(cont + 1)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(cont + 1)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(cont + 1)] = nome_org.value
    Aba["C" + str(cont + 1)] = ano.value
    Aba["D" + str(cont + 1)] = nome_resp.value

    Comb_mov = workbook["Combustão móvel"]

    lista_aux_Comb_mov = leitura_conj_lista(Comb_mov, "D", 207, 356,6)

    '''if lista_aux_Comb_mov == []:
        for i in string.ascii_uppercase[7:13]:
            Aba[i + str(cont+1)] = 0'''

    k=1 #k = 0 omite quem não reportou

    for j, k in zip(lista_aux_Comb_mov, range(1, (len(lista_aux_Comb_mov) + 1))):

        Desc_Frota = Comb_mov["C" + str(j)]
        Tipo_Comb_mov = Comb_mov["D" + str(j)]
        CO2_Foss_cmov = Comb_mov["BB" + str(j)]
        CH4_cmov = Comb_mov["BC" + str(j)]
        N2O_cmov = Comb_mov["BD" + str(j)]
        Emissao_tot_cmov = Comb_mov["BE" + str(j)]
        Emissao_biog_cmov = Comb_mov["BF" + str(j)]

        #Desta forma o usuario pode reportar tanto mes a mes quanto apenas o consumo anual
        if Comb_mov["R" + str(j)].value is None:
            Quant_cons_mov = 0
            for i1 in string.ascii_uppercase[5:17]:
                if Comb_mov[i1 + str(j)].value is not None:
                    Quant_cons_mov = Quant_cons_mov + float(Comb_mov[i1 + str(j)].value)
        else:
            Quant_cons_mov = float(Comb_mov["R" + str(j)].value)

        Aba["E" + str(cont + k)] = Desc_Frota.value
        Aba["F" + str(cont + k)] = Tipo_Comb_mov.value
        Aba["G" + str(cont + k)] = Quant_cons_mov
        Aba["H" + str(cont + k)] = CO2_Foss_cmov.value
        Aba["I" + str(cont + k)] = CH4_cmov.value
        Aba["J" + str(cont + k)] = N2O_cmov.value
        Aba["K" + str(cont + k)] = Emissao_tot_cmov.value
        Aba["L" + str(cont + k)] = Emissao_biog_cmov.value


        '''if Emissao_tot_cmov.value == "0":
            Aba["L" + str(cont + k)] = 0

        if Emissao_biog_cmov.value == "0":
            Aba["M" + str(cont + k)] = 0'''

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

print("Fim do processamento de Combustão Móvel")

#_############################################ Emissoes Fugitivas ################################################

cont = 1
for planilha, i in zip(Plans, range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Emissões fugitivas1"]

    Aba["A" + str(cont + 1)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(cont + 1)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(cont + 1)] = nome_org.value
    Aba["C" + str(cont + 1)] = ano.value
    Aba["D" + str(cont + 1)] = nome_resp.value

    Em_fug = workbook["Emissões fugitivas"]

    lista_aux_Em_fug = leitura_conj_lista(Em_fug, "C", 62, 111,5)

    #if lista_aux_Em_fug == []:
     #   for i in string.ascii_uppercase[6:12]:
      #      Aba[i + str(cont+1)] = 0

    k = 1

    for j, k in zip(lista_aux_Em_fug, range(1, (len(lista_aux_Em_fug,) + 1))):

        Reg_fonte = Em_fug["B" + str(j)]
        gas_ou_comp = Em_fug["C" + str(j)]
        Un_novas_carga = Em_fug["E" + str(j)]
        Un_novas_cap = Em_fug["F" + str(j)]
        Un_exist_rec = Em_fug["G" + str(j)]
        Un_disp_cap = Em_fug["H"+str(j)]
        Un_disp_rec = Em_fug["I"+str(j)]
        Emissao_eq_fug1 = Em_fug["J" + str(j)]

        Aba["E" + str(cont + k)] = Reg_fonte.value
        Aba["F" + str(cont + k)] = gas_ou_comp.value
        Aba["G" + str(cont + k)] = Un_novas_carga.value
        Aba["H" + str(cont + k)] = Un_novas_cap.value
        Aba["I" + str(cont + k)] = Un_exist_rec.value
        Aba["J" + str(cont + k)] = Un_disp_cap.value
        Aba["K" + str(cont + k)] = Un_disp_rec.value
        Aba["L" + str(cont + k)] = Emissao_eq_fug1.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k


cont = 1
for planilha, i in zip(Plans, range(len(Plans))):
    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Emissões fugitivas2"]

    Aba["A" + str(cont + 1)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(cont + 1)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(cont + 1)] = nome_org.value
    Aba["C" + str(cont + 1)] = ano.value
    Aba["D" + str(cont + 1)] = nome_resp.value

    Em_fug = workbook["Emissões fugitivas"]

    lista_aux_Em_fug2 = leitura_conj_lista(Em_fug, "D", 237, 271,5)

    k = 1
    for j, k in zip(lista_aux_Em_fug2, range(1, (len(lista_aux_Em_fug2, ) + 1))):

        Reg_fonte = Em_fug["B" + str(j)]
        Desc_Fonte = Em_fug["C" + str(j)]
        GEE = Em_fug["D" + str(j)]
        Emissoes_GEE = Em_fug["E" + str(j)]
        Emissoes_CO2eq = Em_fug["G" + str(j)]

        Aba["E" + str(cont + k)] = Reg_fonte.value
        Aba["F" + str(cont + k)] = Desc_Fonte.value
        Aba["G" + str(cont + k)] = GEE.value
        Aba["H" + str(cont + k)] = Emissoes_GEE.value
        Aba["I" + str(cont + k)] = Emissoes_CO2eq.value


        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

print("Fim do processamento de Emissões Fugitivas")

#_################################################## Resíduos sólidos ################################################################


for planilha, i in zip(Plans, range(len(Plans))):

    cont = i+1
    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Resíduos sólidos1"]

    Aba["A" + str(cont + 1)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(cont + 1)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(cont + 1)] = nome_org.value
    Aba["C" + str(cont + 1)] = ano.value
    Aba["D" + str(cont + 1)] = nome_resp.value

    Res_sol = workbook["Resíduos sólidos"]

    Massa_res = Res_sol["F" + str(135)]
    Em_CH4_comp = Res_sol["E" + str(154)]
    Em_N2O_comp = Res_sol["E" + str(155)]
    Em_tCO2_comp = Res_sol["E" + str(156)]
    Em_CO2biog_comp = Res_sol["E" + str(157)]

    Aba["E" + str(cont + 1)] = Massa_res.value
    Aba["F" + str(cont + 1)] = Em_CH4_comp.value
    Aba["G" + str(cont + 1)] = Em_N2O_comp.value
    Aba["H" + str(cont + 1)] = Em_tCO2_comp.value
    Aba["I" + str(cont + 1)] = Em_CO2biog_comp.value


cont = 1
for planilha, i in zip(Plans, range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Resíduos sólidos2"]

    Aba["A" + str(cont + 1)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(cont + 1)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(cont + 1)] = nome_org.value
    Aba["C" + str(cont + 1)] = ano.value
    Aba["D" + str(cont + 1)] = nome_resp.value

    Res_sol = workbook["Resíduos sólidos"]

    lista_aux_Res_sol = leitura_conj_lista(Res_sol, "C", 170, 189,5)

    #if lista_aux_Comb_mov == []:
    #   for i in string.ascii_uppercase[7:13]:
    #        Aba[i + str(cont+1)] = 0

    k=1 #k = 0 omite quem não reportou

    for j, k in zip(lista_aux_Res_sol, range(1, (len(lista_aux_Res_sol) + 1))):

        Tipo_res = Res_sol["C" + str(j)]
        Em_CO2_inc = Res_sol["D" + str(j)]
        Em_CH4_inc = Res_sol["E" + str(j)]
        Em_N2O_inc = Res_sol["F" + str(j)]
        Em_CO2eq_inc = Res_sol["G" + str(j)]
        Em_CO2biog_inc = Res_sol["H" + str(j)]

        Aba["E" + str(cont + k)] = Tipo_res.value
        Aba["F" + str(cont + k)] =  Em_CO2_inc.value
        Aba["G" + str(cont + k)] = Em_CH4_inc.value
        Aba["H" + str(cont + k)] = Em_N2O_inc.value
        Aba["I" + str(cont + k)] = Em_CO2eq_inc.value
        Aba["J" + str(cont + k)] = Em_CO2biog_inc.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

print("Fim do processamento de Resíduos Sólidos")

#_############################################## Efluentes ###################################################

for planilha, i in zip(Plans, range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Efluentes"]

    Aba["A" + str(i + 2)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(i + 2)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(i + 2)] = nome_org.value
    Aba["C" + str(i + 2)] = ano.value
    Aba["D" + str(i + 2)] = nome_resp.value

    Efluentes = workbook["Efluentes"]

    Tipo_trat_seq = Efluentes["E" + str(34)]
    Quant_efl_ano = Efluentes["D" + str(39)]
    Comp_org_degr = Efluentes["D" + str(46)]
    Unidade_efl = Efluentes["E" + str(46)]  # Unidade
    Quant_N2_efl = Efluentes["D" + str(52)]  # Quantidade N2
    Fat_em_N2O = Efluentes["D" + str(53)]  # Fator emissão N2O
    Nit_remov_lodo = Efluentes["D" + str(54)]  # N removido
    Tipo_trat = Efluentes["D" + str(58)]  # Tipo de tratamento
    Fat_conv_CH4 = Efluentes["D" + str(59)]  # Fator de conversão de metano
    CH4_rec = Efluentes["E" + str(64)]  # CH4 recuperado
    Dest_biog = Efluentes["D" + str(67)]  # Destino biogas

    Aba["E" + str(i+2)] = Tipo_trat_seq.value
    Aba["F" + str(i+2)] = Quant_efl_ano.value
    Aba["G" + str(i+2)] = Comp_org_degr.value
    Aba["H" + str(i+2)] = Unidade_efl.value
    Aba["I" + str(i+2)] = Quant_N2_efl.value
    Aba["J" + str(i+2)] = Fat_em_N2O.value
    Aba["K" + str(i+2)] = Nit_remov_lodo.value
    Aba["L" + str(i+2)] = Tipo_trat.value
    Aba["M" + str(i+2)] = Fat_conv_CH4.value
    Aba["N" + str(i+2)] = CH4_rec.value
    Aba["O" + str(i+2)] = Dest_biog.value

print("Fim do processamento de Efluentes")

#_########################################################### En. Eletrica (localizacao) ###############################################################################

cont = 1
for planilha,i in zip(Plans,range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only = True, data_only= True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["En. elétrica (localização)1"]

    Aba["A"+str(cont+1)] = int(re.findall('(^.+?)_' ,Plans[i])[0])
    Aba["A" + str(cont+1)].number_format =  numbers.builtin_format_code(1)
    Aba["B"+str(cont+1)] = nome_org.value
    Aba["C"+str(cont+1)] = ano.value
    Aba["D"+str(cont+1)] = nome_resp.value

    En_Elet = workbook["En. elétrica (localização)"]

    lista_aux_En_Elet = leitura_conj_lista(En_Elet, "Q", 41, 90,5)

   # if lista_aux_Comb_est == []:
   #     for i in string.ascii_uppercase[4:15]:
   #         if i != "H":
   #             Aba["letra" + str(cont+1)] = 0

    for j,k in zip(lista_aux_En_Elet, range(1,(len(lista_aux_En_Elet)+1))):

        Desc_Fonte_En_elet = En_Elet["C" + str(j)]
        Elet_Total_comprada = En_Elet["Q" + str(j)]
        Emissao_CO2_elet = En_Elet["AD" + str(j)]

        Aba["E" + str(cont+k)] = Desc_Fonte_En_elet.value
        Aba["F" + str(cont+k)] = Elet_Total_comprada.value
        Aba["G" + str(cont+k)] = Emissao_CO2_elet.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

cont = 1
for planilha,i in zip(Plans,range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only = True, data_only= True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["En. elétrica (localização)2"]

    Aba["A"+str(cont+1)] = int(re.findall('(^.+?)_' ,Plans[i])[0])
    Aba["A" + str(cont+1)].number_format =  numbers.builtin_format_code(1)
    Aba["B"+str(cont+1)] = nome_org.value
    Aba["C"+str(cont+1)] = ano.value
    Aba["D"+str(cont+1)] = nome_resp.value

    En_Elet = workbook["En. elétrica (localização)"]

    lista_aux_En_Elet = leitura_conj_lista(En_Elet, "Q", 106, 124,5)

   # if lista_aux_Comb_est == []:
   #     for i in string.ascii_uppercase[4:15]:
   #         if i != "H":
   #             Aba["letra" + str(cont+1)] = 0

    for j,k in zip(lista_aux_En_Elet, range(1,(len(lista_aux_En_Elet)+1))):

        Desc_Fonte_En_elet = En_Elet["C" + str(j)]
        Elet_Total_comprada = En_Elet["Q" + str(j)]
        Emissao_CO2_elet = En_Elet["AE" + str(j)]

        Aba["E" + str(cont+k)] = Desc_Fonte_En_elet.value
        Aba["F" + str(cont+k)] = Elet_Total_comprada.value
        Aba["G" + str(cont+k)] = Emissao_CO2_elet.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

cont = 1
for planilha,i in zip(Plans,range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only = True, data_only= True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["En. elétrica (localização)3"]

    Aba["A"+str(cont+1)] = int(re.findall('(^.+?)_' ,Plans[i])[0])
    Aba["A" + str(cont+1)].number_format =  numbers.builtin_format_code(1)
    Aba["B"+str(cont+1)] = nome_org.value
    Aba["C"+str(cont+1)] = ano.value
    Aba["D"+str(cont+1)] = nome_resp.value

    En_Elet = workbook["En. elétrica (localização)"]

    lista_aux_En_Elet_aux1 = leitura_conj_lista(En_Elet, "C", 137, 146,5)
    lista_aux_En_Elet_aux2 = leitura_conj_lista(En_Elet, "D", 137, 146,5)
    lista_aux_En_Elet_aux3 = leitura_conj_lista(En_Elet, "F", 137, 146,5)
    lista_aux_En_Elet_aux4 = leitura_conj_lista(En_Elet, "H", 137, 146,5)
    lista_aux_En_Elet_aux5 = leitura_conj_lista(En_Elet, "L", 137, 146,5)

    lista_aux_En_Elet = unique(lista_aux_En_Elet_aux1 + lista_aux_En_Elet_aux2 + lista_aux_En_Elet_aux3 + lista_aux_En_Elet_aux4 + lista_aux_En_Elet_aux5)

    k=1

    for j,k in zip(lista_aux_En_Elet, range(1,(len(lista_aux_En_Elet)+1))):

        Desc_Fonte_En_elet = En_Elet["C" + str(j)]
        Emissao_CO2_elet = En_Elet["D" + str(j)]
        Emissao_CH4_elet = En_Elet["F" + str(j)]
        Emissao_N2O_elet = En_Elet["H" + str(j)]
        Emissao_CO2e_elet = En_Elet["J" + str(j)]
        Emissao_CO2biog_elet = En_Elet["L" + str(j)]

        Aba["E" + str(cont+k)] = Desc_Fonte_En_elet.value
        Aba["F" + str(cont+k)] = Emissao_CO2_elet.value
        Aba["G" + str(cont+k)] = Emissao_CH4_elet.value
        Aba["H" + str(cont + k)] = Emissao_N2O_elet.value
        Aba["I" + str(cont + k)] = Emissao_CO2e_elet.value
        Aba["J" + str(cont + k)] = Emissao_CO2biog_elet.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

print("Fim do processamento de En. Elétrica (localização)")

#_################################### En. elétrica (escolha de compra) ##################################

cont = 1
for planilha,i in zip(Plans,range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only = True, data_only= True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["En. elétrica(escolha de compra)"]

    Aba["A"+str(cont+1)] = int(re.findall('(^.+?)_' ,Plans[i])[0])
    Aba["A" + str(cont+1)].number_format =  numbers.builtin_format_code(1)
    Aba["B"+str(cont+1)] = nome_org.value
    Aba["C"+str(cont+1)] = ano.value
    Aba["D"+str(cont+1)] = nome_resp.value

    En_Elet_esc = workbook["En. elétrica(escolha de compra)"]

    lista_aux_En_Elet_aux1 = leitura_conj_lista(En_Elet_esc, "D", 43, 92,5)

    k=1

    for j,k in zip(lista_aux_En_Elet_aux1, range(1,(len(lista_aux_En_Elet_aux1)+1))):

        Desc_Fonte_esc = En_Elet_esc["C" + str(j)]
        Tipo_font_esc = En_Elet_esc["D" + str(j)]
        Comb_esc = En_Elet_esc["E" + str(j)]
        fat_em_esc = En_Elet_esc["F" + str(j)]
        Ef_planta_esc = En_Elet_esc["G" + str(j)]
        elet_compra_esc = En_Elet_esc["AC" + str(j)]
        emiss_CO2_esc = En_Elet_esc["AD" + str(j)]
        emiss_CH4_esc = En_Elet_esc["AE" + str(j)]
        emiss_N2O_esc = En_Elet_esc["AF" + str(j)]
        emiss_CO2e_esc = En_Elet_esc["AG" + str(j)]
        emiss_CO2biog_esc = En_Elet_esc["AH" + str(j)]

        Aba["E" + str(cont+k)] = Desc_Fonte_esc.value
        Aba["F" + str(cont+k)] = Tipo_font_esc.value
        Aba["G" + str(cont+k)] = Comb_esc.value
        Aba["H" + str(cont + k)] = fat_em_esc.value
        Aba["I" + str(cont + k)] = Ef_planta_esc.value
        Aba["J" + str(cont + k)] = elet_compra_esc.value
        Aba["K" + str(cont + k)] = emiss_CO2_esc.value
        Aba["L" + str(cont + k)] = emiss_CH4_esc.value
        Aba["M" + str(cont + k)] = emiss_N2O_esc.value
        Aba["N" + str(cont + k)] = emiss_CO2e_esc.value
        Aba["O" + str(cont + k)] = emiss_CO2biog_esc.value


        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

print("Fim do processamento de En. Elétrica (escolha de compra)")

#_########################################### Transp.& Distribuição(Upstream) ################################################

cont = 1
for planilha,i in zip(Plans,range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only = True, data_only= True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida.worksheets[cont_aux2]
    T_D_Up = workbook.worksheets[cont_aux]

    Aba["A"+str(cont+1)] = int(re.findall('(^.+?)_' ,Plans[i])[0])
    Aba["A" + str(cont+1)].number_format =  numbers.builtin_format_code(1)
    Aba["B"+str(cont+1)] = nome_org.value
    Aba["C"+str(cont+1)] = ano.value
    Aba["D"+str(cont+1)] = nome_resp.value

    lista_aux_T_D_Up = leitura_conj_lista(T_D_Up, "L", 540, 689,5)

    k=1

    for j,k in zip(lista_aux_T_D_Up , range(1,(len(lista_aux_T_D_Up)+1))):

        Desc_Fonte_T_D = T_D_Up["C" + str(j)]
        Tipo_Frota_T_D = T_D_Up["L" + str(j)]
        Ano_Frota_T_D = T_D_Up["M" + str(j)]

        #Desta forma o usuario pode reportar tanto mes a mes quanto apenas a distância anual
        if T_D_Up["AH" + str(j)].value is None:
            Dist_T_D = 0
            for i1 in ["V","W","X","Y","Z","AA","AB", "AC", "AD", "AE", "AF", "AG"]:
                if T_D_Up[i1 + str(j)].value is not None:
                    Dist_T_D = Dist_T_D + T_D_Up[i1 + str(j)].value
        else:
            Dist_T_D = T_D_Up["AH" + str(j)].value

        Emiss_CO2_T_D = T_D_Up["CE" + str(j)]
        Emiss_CH4_T_D = T_D_Up["CF" + str(j)]
        Emiss_N2O_T_D = T_D_Up["CG" + str(j)]
        Emiss_tot_T_D = T_D_Up["CH" + str(j)]
        Emiss_CO2biog_T_D = T_D_Up["CI" + str(j)]

        Aba["E" + str(cont + k)] = Desc_Fonte_T_D.value
        Aba["F" + str(cont + k)] = Tipo_Frota_T_D.value
        Aba["G" + str(cont + k)] = Ano_Frota_T_D.value
        Aba["H" + str(cont + k)] = Dist_T_D
        Aba["I" + str(cont + k)] = Emiss_CO2_T_D.value
        Aba["J" + str(cont + k)] = Emiss_CH4_T_D.value
        Aba["K" + str(cont + k)] = Emiss_N2O_T_D.value
        Aba["L" + str(cont + k)] = Emiss_tot_T_D.value
        Aba["M" + str(cont + k)] = Emiss_CO2biog_T_D.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

print("Fim do processamento de Transp.& Distribuição(Upstream)")


#_#################################################### Resíduos Sólidos da Operação1 #################################################
cont = 1
for planilha,i in zip(Plans,range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only = True, data_only= True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Resíduos sólidos da operação1"]
    Res_sol_op = workbook["Resíduos sólidos da operação"]

    Aba["A"+str(cont+1)] = int(re.findall('(^.+?)_' ,Plans[i])[0])
    Aba["A" + str(cont+1)].number_format =  numbers.builtin_format_code(1)
    Aba["B"+str(cont+1)] = nome_org.value
    Aba["C"+str(cont+1)] = ano.value
    Aba["D"+str(cont+1)] = nome_resp.value

    Aba["E" + str(cont+1)] = Res_sol_op["D" + str(54)].value
    Aba["F" + str(cont + 1)] = Res_sol_op["D" + str(55)].value
    Aba["G" + str(cont + 1)] = Res_sol_op["E" + str(67)].value
    Aba["H" + str(cont + 1)] = Res_sol_op["E" + str(103)].value
    Aba["I" + str(cont + 1)] = Res_sol_op["E" + str(114)].value
    Aba["J" + str(cont + 1)] = Res_sol_op["E" + str(125)].value
    Aba["K" + str(cont + 1)] = Res_sol_op["E" + str(126)].value
    Aba["L" + str(cont + 1)] = Res_sol_op["E" + str(127)].value

    cont = cont + 1


#_################################ Resíduos Sólidos da Operação2 ###################################################

cont = 1
for planilha,i in zip(Plans,range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only = True, data_only= True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Resíduos sólidos da operação2"]
    Res_sol_op = workbook["Resíduos sólidos da operação"]

    Aba["A"+str(cont+1)] = int(re.findall('(^.+?)_' ,Plans[i])[0])
    Aba["A" + str(cont+1)].number_format =  numbers.builtin_format_code(1)
    Aba["B"+str(cont+1)] = nome_org.value
    Aba["C"+str(cont+1)] = ano.value
    Aba["D"+str(cont+1)] = nome_resp.value

    lista_aux_Res_sol_op = leitura_conj_lista(Res_sol_op, "C", 169, 188,5)

    k=1

    for j,k in zip(lista_aux_Res_sol_op, range(1,(len(lista_aux_Res_sol_op)+1))):

        Tipo_res = Res_sol_op["C" + str(j)]
        Em_CO2_inc_3 = Res_sol_op["D" + str(j)]
        Em_CH4_inc_3 = Res_sol_op["E" + str(j)]
        Em_N2O_inc_3 = Res_sol_op["F" + str(j)]
        Em_CO2e_inc_3 = Res_sol_op["G" + str(j)]
        Em_CO2biog_inc_3 = Res_sol_op["H" + str(j)]

        Aba["E" + str(cont + k)] = Tipo_res.value
        Aba["F" + str(cont + k)] = Em_CO2_inc_3.value
        Aba["G" + str(cont + k)] = Em_CH4_inc_3.value
        Aba["H" + str(cont + k)] = Em_N2O_inc_3.value
        Aba["I" + str(cont + k)] = Em_CO2e_inc_3.value
        Aba["J" + str(cont + k)] = Em_CO2biog_inc_3.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

print("Fim do processamento de Resíduos Sólidos da Operação")

#_################################# Efluentes gerados na operação #################################################

cont = 1
for planilha,i in zip(Plans,range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only = True, data_only= True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Efluentes gerados na operação"]
    Eflu_op = workbook["Efluentes gerados na operação"]

    Aba["A"+str(cont+1)] = int(re.findall('(^.+?)_' ,Plans[i])[0])
    Aba["A" + str(cont+1)].number_format =  numbers.builtin_format_code(1)
    Aba["B"+str(cont+1)] = nome_org.value
    Aba["C"+str(cont+1)] = ano.value
    Aba["D"+str(cont+1)] = nome_resp.value

    Trat_eflu = Eflu_op["E" + str(40)]
    Quant_eflu = Eflu_op["D" + str(45)]
    Comp_org = Eflu_op["D" + str(52)]
    Unidade_eflu_3_2 = Eflu_op["E" + str(52)]

    Quant_N = Eflu_op["D" + str(58)]
    Fat_em = Eflu_op["D" + str(59)]
    Nit_remov = Eflu_op["D" + str(60)]
    Tipo_trat_eflu = Eflu_op["D" + str(64)]
    Fat_conv_CH4_eflu = Eflu_op["D" + str(65)]
    Quant_CH4_rec = Eflu_op["E" + str(70)]
    Quant_eflu_ap = Eflu_op["D" + str(79)]
    Comp_org_degr_eflu = Eflu_op["D" + str(86)]
    Unidade_eflu_3 = Eflu_op["E" + str(86)]
    Quant_N_eflu_3 = Eflu_op["D" + str(92)]
    Fat_N_eflu_3 = Eflu_op["D" + str(93)]
    lodo_N_eflu_3 = Eflu_op["D" + str(94)]

    Tipo_trat_eflu_2 = Eflu_op["D" + str(98)]
    Fat_conv_CH4_eflu_2 = Eflu_op["D" + str(99)]
    CH4_rec_eflu_3 = Eflu_op["D" + str(104)]
    CH4_trat_eflu_3 = Eflu_op["E" + str(110)]
    N2O_trat_eflu_3 = Eflu_op["E" + str(111)]
    CO2_trat_eflu_3 = Eflu_op["E" + str(112)]
    CO2biog_trat_eflu_3 = Eflu_op["E" + str(113)]

    Aba["E" + str(cont + 1)] = Trat_eflu.value
    Aba["F" + str(cont + 1)] = Quant_eflu.value
    Aba["G" + str(cont + 1)] = Comp_org.value
    Aba["H" + str(cont + 1)] = Unidade_eflu_3_2.value

    Aba["I" + str(cont + 1)] = Quant_N.value
    Aba["J" + str(cont + 1)] = Fat_em.value
    Aba["K" + str(cont + 1)] = Nit_remov.value
    Aba["L" + str(cont + 1)] = Tipo_trat_eflu.value
    Aba["M" + str(cont + 1)] = Fat_conv_CH4_eflu.value
    Aba["N" + str(cont + 1)] = Quant_CH4_rec.value
    Aba["O" + str(cont + 1)] = Quant_eflu_ap.value
    Aba["P" + str(cont + 1)] = Comp_org_degr_eflu.value
    Aba["Q" + str(cont + 1)] = Unidade_eflu_3.value
    Aba["R" + str(cont + 1)] = Quant_N_eflu_3.value
    Aba["S" + str(cont + 1)] = Fat_N_eflu_3.value
    Aba["T" + str(cont + 1)] = lodo_N_eflu_3.value

    Aba["U" + str(cont + 1)] = Tipo_trat_eflu_2.value
    Aba["V" + str(cont + 1)] = Fat_conv_CH4_eflu_2.value
    Aba["W" + str(cont + 1)] = CH4_rec_eflu_3.value
    Aba["X" + str(cont + 1)] = CH4_trat_eflu_3.value
    Aba["Y" + str(cont + 1)] = N2O_trat_eflu_3.value
    Aba["Z" + str(cont + 1)] = CO2_trat_eflu_3.value
    Aba["AA" + str(cont + 1)] = CO2biog_trat_eflu_3.value

    cont = cont + 1

print("Fim do processamento de Efluentes gerados na operação")

#_######################################## Viagens a Negócios1 ###################################

cont = 1
for planilha, i in zip(Plans, range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Viagens a Negócios1"]

    Aba["A" + str(cont + 1)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(cont + 1)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(cont + 1)] = nome_org.value
    Aba["C" + str(cont + 1)] = ano.value
    Aba["D" + str(cont + 1)] = nome_resp.value

    Viagens = workbook["Viagens a Negócios"]

    lista_aux_Viagens1 = leitura_conj_lista(Viagens, "W", 47, 796,5)

    k=1 #k = 0 omite quem não reportou

    for j, k in zip(lista_aux_Viagens1, range(1, (len(lista_aux_Viagens1) + 1))):

        Dist_trech = Viagens["W" + str(j)]
        Nr_trech = Viagens["X" + str(j)]
        Dist_total = Viagens["Y" + str(j)]
        Em_CO2_viagens1 = Viagens["AC" + str(j)]
        Em_CH4_viagens1 = Viagens["AD" + str(j)]
        Em_N2O_viagens1 = Viagens["AE" + str(j)]
        Em_CO2_tot_viagens1 = Viagens["AF" + str(j)]

        Aba["E" + str(cont + k)] = Dist_trech.value
        Aba["F" + str(cont + k)] = Nr_trech.value
        Aba["G" + str(cont + k)] = Dist_total.value
        Aba["H" + str(cont + k)] = Em_CO2_viagens1.value
        Aba["I" + str(cont + k)] = Em_CH4_viagens1.value
        Aba["J" + str(cont + k)] = Em_N2O_viagens1.value
        Aba["K" + str(cont + k)] = Em_CO2_tot_viagens1.value
        


        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k



#_######################################## Viagens a Negócios2 ###################################

cont = 1
for planilha, i in zip(Plans, range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Viagens a Negócios2"]

    Aba["A" + str(cont + 1)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(cont + 1)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(cont + 1)] = nome_org.value
    Aba["C" + str(cont + 1)] = ano.value
    Aba["D" + str(cont + 1)] = nome_resp.value

    Viagens = workbook["Viagens a Negócios"]

    lista_aux_Viagens = leitura_conj_lista(Viagens, "L", 1102, 1131,5)

    k=1 #k = 0 omite quem não reportou

    for j, k in zip(lista_aux_Viagens, range(1, (len(lista_aux_Viagens) + 1))):

        Desc_pass= Viagens["C" + str(j)]
        Tipo_frota_Viagens = Viagens["L" + str(j)]
        Ano_frota_Viagens = Viagens["M" + str(j)]

        #Desta forma o usuario pode reportar tanto mes a mes quanto apenas a distância anual
        if Viagens["AH" + str(j)].value is None:
            Dist_anual_Viagens = 0
            for i1 in range(22,34):
                char = lista_str_num(i1)
                if Viagens[char + str(j)].value is not None:
                    Dist_anual_Viagens = Dist_anual_Viagens + Viagens[char + str(j)].value
        else:
            Dist_anual_Viagens = Viagens["AH" + str(j)].value

        Em_CO2_viagens = Viagens["CE" + str(j)]
        Em_CH4_viagens = Viagens["CF" + str(j)]
        Em_N2O_viagens = Viagens["CG" + str(j)]
        Em_CO2_tot_viagens = Viagens["CH" + str(j)]
        Em_CO2_biog_viagens = Viagens["CI" + str(j)]

        Aba["E" + str(cont + k)] = Desc_pass.value
        Aba["F" + str(cont + k)] = Tipo_frota_Viagens.value
        Aba["G" + str(cont + k)] = Ano_frota_Viagens.value
        Aba["H" + str(cont + k)] = Dist_anual_Viagens
        Aba["I" + str(cont + k)] = Em_CO2_viagens.value
        Aba["J" + str(cont + k)] = Em_CH4_viagens.value
        Aba["K" + str(cont + k)] = Em_N2O_viagens.value
        Aba["L" + str(cont + k)] = Em_CO2_tot_viagens.value
        Aba["M" + str(cont + k)] = Em_CO2_biog_viagens.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

print("Fim do processamento de Viagens a Negócios")


#_################################### Deslocamento casa-trabalho ##############################################
#1

cont = 1
for planilha, i in zip(Plans, range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Deslocamento casa-trabalho1"]

    Aba["A" + str(cont + 1)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(cont + 1)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(cont + 1)] = nome_org.value
    Aba["C" + str(cont + 1)] = ano.value
    Aba["D" + str(cont + 1)] = nome_resp.value

    Desloc_CT = workbook["Deslocamento casa-trabalho"]

    lista_aux_Desloc_CT = leitura_conj_lista(Desloc_CT, "D", 40, 138,5)

    k=1 #k = 0 omite quem não reportou

    for j, k in zip(lista_aux_Desloc_CT, range(1, (len(lista_aux_Desloc_CT) + 1))):

        Tipo_transp = Desloc_CT["D" + str(j)]
        Dist_ano = Desloc_CT["F" + str(j)]
        Dias_trab_ano = Desloc_CT["G" + str(j)]
        Em_CO2_transp = Desloc_CT["K" + str(j)]
        Em_CO2_tot = Desloc_CT["N" + str(j)]

        Aba["E" + str(cont + k)] = Tipo_transp.value
        Aba["F" + str(cont + k)] = Dist_ano.value
        Aba["G" + str(cont + k)] = Dias_trab_ano.value
        Aba["H" + str(cont + k)] = Em_CO2_transp.value
        Aba["I" + str(cont + k)] = Em_CO2_tot.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

#2

cont = 1
for planilha, i in zip(Plans, range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Deslocamento casa-trabalho2"]

    Aba["A" + str(cont + 1)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(cont + 1)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(cont + 1)] = nome_org.value
    Aba["C" + str(cont + 1)] = ano.value
    Aba["D" + str(cont + 1)] = nome_resp.value

    Desloc_CT = workbook["Deslocamento casa-trabalho"]

    lista_aux_Desloc_CT2 = leitura_conj_lista(Desloc_CT, "D", 152, 251,5)

    k = 1  # k = 0 omite quem não reportou

    for j, k in zip(lista_aux_Desloc_CT2, range(1, (len(lista_aux_Desloc_CT2) + 1))):

        Tipo_onib = Desloc_CT["D" + str(j)]
        Dist_ano_onib = Desloc_CT["F" + str(j)]
        Dias_trab_ano_onib = Desloc_CT["G" + str(j)]
        Em_CO2_onib = Desloc_CT["W" + str(j)]
        Em_CO2_tot_onib = Desloc_CT["X" + str(j)]

        Aba["E" + str(cont + k)] = Tipo_onib.value
        Aba["F" + str(cont + k)] = Dist_ano_onib.value
        Aba["G" + str(cont + k)] = Dias_trab_ano_onib.value
        Aba["H" + str(cont + k)] = Em_CO2_onib.value
        Aba["I" + str(cont + k)] = Em_CO2_tot_onib.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

#3

cont = 1
for planilha, i in zip(Plans, range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Deslocamento casa-trabalho3"]

    Aba["A" + str(cont + 1)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(cont + 1)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(cont + 1)] = nome_org.value
    Aba["C" + str(cont + 1)] = ano.value
    Aba["D" + str(cont + 1)] = nome_resp.value

    Desloc_CT = workbook["Deslocamento casa-trabalho"]

    lista_aux_Desloc_CT3 = leitura_conj_lista(Desloc_CT, "D", 265, 334,5)

    k = 1  # k = 0 omite quem não reportou

    for j, k in zip(lista_aux_Desloc_CT3, range(1, (len(lista_aux_Desloc_CT3) + 1))):
        Tipo_balsa = Desloc_CT["D" + str(j)]
        Dist_ano_balsa = Desloc_CT["F" + str(j)]
        Dias_trab_ano_balsa = Desloc_CT["G" + str(j)]
        Em_CO2_balsa = Desloc_CT["K" + str(j)]
        Em_CH4_balsa = Desloc_CT["L" + str(j)]
        Em_N2O_balsa = Desloc_CT["M" + str(j)]
        Em_CO2_tot_balsa = Desloc_CT["N" + str(j)]

        Aba["E" + str(cont + k)] = Tipo_balsa.value
        Aba["F" + str(cont + k)] = Dist_ano_balsa.value
        Aba["G" + str(cont + k)] = Dias_trab_ano_balsa.value
        Aba["H" + str(cont + k)] = Em_CO2_balsa.value
        Aba["I" + str(cont + k)] = Em_CH4_balsa.value
        Aba["J" + str(cont + k)] = Em_N2O_balsa.value
        Aba["K" + str(cont + k)] = Em_CO2_tot_balsa.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

#4
cont = 1
for planilha, i in zip(Plans, range(len(Plans))):

    endereco = pasta_BD + "\\" + planilha
    workbook = load_workbook(endereco, read_only=True, data_only=True)

    intro = workbook["Introdução"]

    nome_org = intro["E" + str(22)]
    ano = intro["E" + str(25)]
    nome_resp = intro["E" + str(27)]

    Aba = plan_saida["Deslocamento casa-trabalho4"]

    Aba["A" + str(cont + 1)] = int(re.findall('(^.+?)_', Plans[i])[0])
    Aba["A" + str(cont + 1)].number_format = numbers.builtin_format_code(1)
    Aba["B" + str(cont + 1)] = nome_org.value
    Aba["C" + str(cont + 1)] = ano.value
    Aba["D" + str(cont + 1)] = nome_resp.value

    Desloc_CT = workbook["Deslocamento casa-trabalho"]

    lista_aux_Desloc_CT4 = leitura_conj_lista(Desloc_CT, "D", 472, 571,5)

    k = 1  # k = 0 omite quem não reportou

    for j, k in zip(lista_aux_Desloc_CT4, range(1, (len(lista_aux_Desloc_CT4) + 1))):

        Tipo_comb = Desloc_CT["D" + str(j)]
        Cons_med = Desloc_CT["F" + str(j)]

        #Desta forma o usuario pode reportar tanto mes a mes quanto apenas os dias trabalhados no ano
        if Desloc_CT["T" + str(j)].value is None:
            Dias_trab_ano_vp = 0
            for i1 in range(8,20):
                char = lista_str_num(i1)
                if Desloc_CT[char + str(j)].value is not None:
                    Dias_trab_ano_vp = Dias_trab_ano_vp + Desloc_CT[char + str(j)].value
        else:
            Dias_trab_ano_vp = Desloc_CT["T" + str(j)].value

        Em_CO2_vp = Desloc_CT["BC" + str(j)]
        Em_CH4_vp = Desloc_CT["BD" + str(j)]
        Em_N2O_vp = Desloc_CT["BE" + str(j)]
        Em_CO2_tot_vp = Desloc_CT["BF" + str(j)]
        Em_CO2_biog_vp = Desloc_CT["BG" + str(j)]

        Aba["E" + str(cont + k)] = Tipo_comb.value
        Aba["F" + str(cont + k)] = Cons_med.value
        Aba["G" + str(cont + k)] = Dias_trab_ano_vp
        Aba["H" + str(cont + k)] = Em_CO2_vp.value
        Aba["I" + str(cont + k)] = Em_CH4_vp.value
        Aba["J" + str(cont + k)] = Em_N2O_vp.value
        Aba["K" + str(cont + k)] = Em_CO2_tot_vp.value
        Aba["L" + str(cont + k)] = Em_CO2_biog_vp.value

        Aba["A" + str(cont + k)] = int(re.findall('(^.+?)_', Plans[i])[0])
        Aba["A" + str(cont + k)].number_format = numbers.builtin_format_code(1)
        Aba["B" + str(cont + k)] = nome_org.value
        Aba["C" + str(cont + k)] = ano.value
        Aba["D" + str(cont + k)] = nome_resp.value

    cont = cont + k

print("Fim do processamento de Deslocamento Casa Trabalho")

plan_saida.save(filename=filename)

print(time.strftime("%H:%M:%S"))

##



